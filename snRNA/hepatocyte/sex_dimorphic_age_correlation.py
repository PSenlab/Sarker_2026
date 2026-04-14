#!/usr/bin/env python3
# ==============================================================================
# Sex-Dimorphic Age-Associated Transcriptional Correlation Analysis
# ==============================================================================
#
# Description:
#   Computes Spearman correlations between gene expression and ordered age
#   groups separately for each sex in hepatocytes, classifies genes into
#   sex-specific aging patterns (Male Only, Female Only, Both Up/Down,
#   Reversal), and runs Reactome pathway enrichment on each group.
#
# Input:
#   - Annotated AnnData (.h5ad) with celltype, sex, and age labels
#
# Output:
#   Part 1 - Age correlation:
#     - Per-sex correlation CSVs (all + significant genes)
#     - Volcano plots per sex
#     - Male vs female scatter plot
#
#   Part 2 - Sex comparison:
#     - Gene classification by sex-specific pattern
#     - Filtered scatter plots with group counts
#     - Highlight plots for reversal patterns
#     - Gene lists per group (CSV)
#
#   Part 3 - Pathway enrichment:
#     - Excel workbook with all Reactome pathways per group
#     - Excel workbook with mouse/human gene lists per group
#     - Bubble plots per group
#     - Combined comparison plot
#
#
# ==============================================================================

import os
import numpy as np
import pandas as pd
import scanpy as sc
import matplotlib.pyplot as plt
import matplotlib as mpl
import seaborn as sns
import mygene
import gseapy as gp
from scipy.stats import spearmanr
from statsmodels.stats.multitest import multipletests
from matplotlib.colors import Normalize
from matplotlib.cm import ScalarMappable
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import warnings

warnings.filterwarnings("ignore")

# Global plot settings
mpl.rcParams["font.family"] = "Arial"
mpl.rcParams["pdf.fonttype"] = 42
mpl.rcParams["ps.fonttype"] = 42


# ==============================================================================
# CONFIGURATION - UPDATE THESE PATHS
# ==============================================================================

INPUT_H5AD = "integrated_scvi.h5ad"
OUTPUT_DIR = "age_corr_results"
COMPARISON_DIR = os.path.join(OUTPUT_DIR, "sex_comparison")
PATHWAY_DIR = os.path.join(OUTPUT_DIR, "pathway_enrichment")

# Thresholds
FDR_THRESHOLD = 0.05
MIN_EXPRESSING_CELLS = 10
CORR_THRESHOLD = 0.05
EXPRESSION_THRESHOLD = 0.05

# Age group ordering (young to geriatric)
AGE_MAPPING = {
    "young": 1,
    "mid_age": 2,
    "old": 3,
    "pre_geriatric": 4,
    "geriatric": 5,
}

# Color palettes
VOLCANO_COLORS = {
    "up": "#E74C3C",
    "down": "#1F77B4",
    "ns": "gray",
}

GROUP_COLORS = {
    "Not Sig": "#D5D8DC",
    "Male Only": "#8E44AD",
    "Female Only": "#C0392B",
    "Both Up": "#2ECC71",
    "Both Down": "#16A085",
    "Reversal M_up_F_down": "skyblue",
    "Reversal M_down_F_up": "lightpink",
}

# Create output directories
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(COMPARISON_DIR, exist_ok=True)
os.makedirs(PATHWAY_DIR, exist_ok=True)


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def compute_age_correlation(adata_subset, sex_name):
    """
    Compute Spearman correlation between gene expression and ordered age groups.

    Returns (df_sig, df_all): significant genes only and all tested genes.
    """
    print(f"\n  Computing age correlations: {sex_name.upper()}")

    age_numeric = adata_subset.obs["age"].map(AGE_MAPPING).astype(float).values
    X = adata_subset.X
    is_sparse = hasattr(X, "toarray")

    print(f"  Cells: {adata_subset.n_obs}, Genes: {adata_subset.n_vars}")

    results = []
    for idx, gene in enumerate(adata_subset.var_names):
        if is_sparse:
            expr = X[:, idx].toarray().flatten()
        else:
            expr = np.asarray(X[:, idx]).flatten()

        if (expr > 0).sum() < MIN_EXPRESSING_CELLS:
            continue

        corr, pval = spearmanr(age_numeric, expr)
        results.append({
            "gene": gene,
            "correlation": corr,
            "pvalue": pval,
            "mean_expression": expr.mean(),
            "pct_expressing": (expr > 0).mean() * 100,
            "sex": sex_name,
        })

    df = pd.DataFrame(results)
    df["pval_adj"] = multipletests(df["pvalue"], method="fdr_bh")[1]
    df["abs_corr"] = df["correlation"].abs()
    df["-log10_fdr"] = -np.log10(df["pval_adj"] + 1e-300)
    df = df.sort_values("abs_corr", ascending=False)

    df_sig = df[df["pval_adj"] < FDR_THRESHOLD].copy()
    n_up = (df_sig["correlation"] > 0).sum()
    n_down = (df_sig["correlation"] < 0).sum()

    print(f"  Genes tested: {len(df)}")
    print(f"  Significant (FDR < {FDR_THRESHOLD}): {len(df_sig)}")
    print(f"    Increase with age: {n_up}")
    print(f"    Decrease with age: {n_down}")

    return df_sig, df


def classify_gene_group(row):
    """
    Classify genes based on sex-specific aging patterns.

    Categories:
    - Reversal M_up_F_down: Positive in male, negative in female
    - Reversal M_down_F_up: Negative in male, positive in female
    - Both Up: Significant in both, positive in both
    - Both Down: Significant in both, negative in both
    - Male Only: Significant only in males
    - Female Only: Significant only in females
    - Not Sig: Not significant in either sex
    """
    m_corr = row["correlation_male"]
    f_corr = row["correlation_female"]
    m_sig = row["sig_male"]
    f_sig = row["sig_female"]

    if m_corr > 0 and f_corr < 0:
        return "Reversal M_up_F_down"
    if m_corr < 0 and f_corr > 0:
        return "Reversal M_down_F_up"

    if m_sig and f_sig:
        if m_corr > 0 and f_corr > 0:
            return "Both Up"
        if m_corr < 0 and f_corr < 0:
            return "Both Down"

    if m_sig and not f_sig:
        return "Male Only"
    if f_sig and not m_sig:
        return "Female Only"

    return "Not Sig"


def get_top_genes(df, n=10):
    """Extract top N positive and negative correlated genes."""
    df_sig = df[df["pval_adj"] < FDR_THRESHOLD].copy()
    top_pos = df_sig.nlargest(n, "correlation")
    top_neg = df_sig.nsmallest(n, "correlation")
    return pd.concat([top_pos, top_neg])


def plot_volcano(df, sex_name, ax):
    """Create volcano plot with labeled top genes."""
    df = df.copy()

    colors = []
    for _, row in df.iterrows():
        if row["pval_adj"] >= FDR_THRESHOLD:
            colors.append(VOLCANO_COLORS["ns"])
        elif row["correlation"] > 0:
            colors.append(VOLCANO_COLORS["up"])
        else:
            colors.append(VOLCANO_COLORS["down"])

    ax.scatter(
        df["correlation"], df["-log10_fdr"],
        c=colors, s=18, alpha=0.7, edgecolors="none",
    )

    ax.axvline(0, color="black", linestyle="--", linewidth=1)
    fdr_line = -np.log10(FDR_THRESHOLD)
    ax.axhline(fdr_line, color="gray", linestyle=":", linewidth=0.8, alpha=0.7)

    top_genes = get_top_genes(df)
    for _, row in top_genes.iterrows():
        ax.annotate(
            row["gene"],
            xy=(row["correlation"], row["-log10_fdr"]),
            fontsize=8,
            ha="left" if row["correlation"] > 0 else "right",
            va="bottom",
        )

    ax.set_title(f"{sex_name.capitalize()} Hepatocytes", fontsize=14, fontweight="bold")
    ax.set_xlabel("Spearman Correlation (Age)", fontsize=12)
    ax.set_ylabel("-log10(FDR)", fontsize=12)
    sns.despine(ax=ax)


def plot_highlight(data, highlight_group, highlight_color, filepath):
    """Create scatter plot highlighting one group."""
    data = data.copy()
    data["highlight"] = data["group"].apply(
        lambda x: highlight_group if x == highlight_group else "Other"
    )

    plt.figure(figsize=(5, 4))

    sns.scatterplot(
        data=data[data["highlight"] == "Other"],
        x="correlation_male", y="correlation_female",
        color="lightgrey", s=40, alpha=0.5, legend=False,
    )
    sns.scatterplot(
        data=data[data["highlight"] == highlight_group],
        x="correlation_male", y="correlation_female",
        color=highlight_color, s=40, alpha=0.9, legend=False,
    )

    plt.axvline(0, linestyle="--", color="black")
    plt.axhline(0, linestyle="--", color="black")
    plt.xlabel("Male Age Correlation", fontsize=12)
    plt.ylabel("Female Age Correlation", fontsize=12)
    plt.title(f"Highlighted: {highlight_group}", fontsize=12, weight="bold")
    plt.tight_layout()
    plt.savefig(filepath, dpi=600, bbox_inches="tight")
    plt.close()
    print(f"  [OK] Saved: {os.path.basename(filepath)}")


def convert_mouse_to_human(mouse_genes):
    """Convert mouse gene symbols to human orthologs via HomoloGene."""
    if not mouse_genes:
        return []

    mg = mygene.MyGeneInfo()

    mouse_to_human = mg.querymany(
        mouse_genes, scopes="symbol", fields="homologene",
        species="mouse", as_dataframe=False, verbose=False,
    )

    human_entrez_ids = []
    for entry in mouse_to_human:
        if isinstance(entry, dict) and "homologene" in entry:
            for homolog in entry["homologene"]["genes"]:
                if homolog[0] == 9606:
                    human_entrez_ids.append(homolog[1])

    human_entrez_ids = list(set(human_entrez_ids))
    if not human_entrez_ids:
        return []

    human_symbol_info = mg.querymany(
        human_entrez_ids, scopes="entrezgene", fields="symbol",
        species="human", verbose=False,
    )

    human_symbols = list({
        entry["symbol"].upper()
        for entry in human_symbol_info
        if isinstance(entry, dict) and "symbol" in entry
    })

    return human_symbols


def run_enrichr_reactome(gene_list, background_genes):
    """Run Enrichr for Reactome 2024 and return full results."""
    if len(gene_list) < 3:
        return pd.DataFrame()

    try:
        enrich_results = gp.enrichr(
            gene_list=gene_list,
            gene_sets="Reactome_Pathways_2024",
            organism="Human",
            background=background_genes,
            outdir=None,
        )

        df = pd.DataFrame(enrich_results.results)
        if df.empty:
            return pd.DataFrame()

        df.columns = df.columns.str.strip()
        df["-log10(P-value)"] = -np.log10(df["P-value"].replace(0, 1e-300))
        df["-log10(Adj P-value)"] = -np.log10(df["Adjusted P-value"].replace(0, 1e-300))

        input_set = set(gene_list)

        def count_overlap(gene_str):
            if not isinstance(gene_str, str):
                return 0
            genes = {g.strip().upper() for g in gene_str.split(";")}
            return len(genes & input_set)

        def list_overlap(gene_str):
            if not isinstance(gene_str, str):
                return ""
            genes = {g.strip().upper() for g in gene_str.split(";")}
            overlap = genes & input_set
            return ";".join(sorted(overlap))

        df["Gene_Count"] = df["Genes"].apply(count_overlap)
        df["Overlap_Genes"] = df["Genes"].apply(list_overlap)

        return df

    except Exception as e:
        print(f"    [ERROR] {e}")
        return pd.DataFrame()


def format_excel(filepath):
    """Apply professional formatting to Excel file."""
    wb = load_workbook(filepath)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

        ws.freeze_panes = "A2"

    wb.save(filepath)


def create_bubble_plot(df, group_name, output_dir, top_n=20):
    """Create publication-quality bubble plot for enrichment results."""
    if df.empty:
        return

    df_sig = df[(df["Adjusted P-value"] < 0.05) & (df["Gene_Count"] >= 3)].copy()
    if df_sig.empty:
        print(f"    No significant pathways for {group_name}")
        return

    top_df = df_sig.sort_values("Odds Ratio", ascending=False).head(top_n)
    top_df = top_df.sort_values("Odds Ratio", ascending=True)
    top_df["Term"] = top_df["Term"].astype(str)

    n_terms = len(top_df)
    fig_height = max(4, n_terms * 0.35)

    plt.figure(figsize=(6, fig_height))
    ax = plt.gca()

    ax.hlines(
        y=top_df["Term"], xmin=0, xmax=top_df["Odds Ratio"],
        color="gray", linewidth=1.2, linestyles="dotted",
    )

    bubble = sns.scatterplot(
        data=top_df, x="Odds Ratio", y="Term",
        size="Gene_Count", hue="Adjusted P-value",
        sizes=(60, 400), palette="coolwarm_r",
        edgecolor="black", linewidth=0.7, alpha=0.95, ax=ax,
    )

    ax.set_title(
        f"Reactome Pathways - {group_name}\n(Top {n_terms}, Adj P < 0.05)",
        fontsize=13, weight="bold",
    )
    ax.set_xlabel("Odds Ratio", fontsize=12, weight="bold")
    ax.set_ylabel("")

    norm = Normalize(
        vmin=top_df["Adjusted P-value"].min(),
        vmax=top_df["Adjusted P-value"].max(),
    )
    sm = ScalarMappable(norm=norm, cmap="coolwarm_r")
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax)
    cbar.set_label("Adjusted P-value", fontsize=11, weight="bold")

    handles, labels = bubble.get_legend_handles_labels()
    ax.legend(
        handles[1:], labels[1:], title="Gene Count",
        bbox_to_anchor=(1.5, 0.7), loc="upper left",
    )

    plt.tight_layout()
    plt.savefig(
        os.path.join(output_dir, f"{group_name}_Reactome_Bubble.pdf"),
        bbox_inches="tight", dpi=300,
    )
    plt.savefig(
        os.path.join(output_dir, f"{group_name}_Reactome_Bubble.png"),
        bbox_inches="tight", dpi=300,
    )
    plt.close()
    print(f"    [OK] Saved: {group_name}_Reactome_Bubble.pdf/png")


# ==============================================================================
# PART 1: AGE CORRELATION ANALYSIS
# ==============================================================================

print()
print("=" * 70)
print("PART 1: Age correlation analysis")
print("=" * 70)

# Step 1: Load and subset data
print("\n  Step 1: Loading data...")
adata = sc.read_h5ad(INPUT_H5AD)
print(f"  Total cells: {adata.n_obs}, Total genes: {adata.n_vars}")

adata_hep = adata[adata.obs["celltype"] == "Hepatocyte"].copy()
adata_male = adata_hep[adata_hep.obs["sex"] == "male"].copy()
adata_female = adata_hep[adata_hep.obs["sex"] == "female"].copy()

print(f"  Hepatocyte cells: {adata_hep.n_obs}")
print(f"  Male: {adata_male.n_obs}, Female: {adata_female.n_obs}")

# Step 2: Run correlation
print("\n  Step 2: Running age correlation...")
male_sig, male_all = compute_age_correlation(adata_male, "male")
female_sig, female_all = compute_age_correlation(adata_female, "female")

# Step 3: Save results
print("\n  Step 3: Saving correlation results...")

def save_results(df, filename):
    filepath = os.path.join(OUTPUT_DIR, f"{filename}.csv")
    df.to_csv(filepath, index=False)
    print(f"  [OK] {filepath}")

save_results(male_sig, "male_hepatocytes_ageCorr_sig")
save_results(male_all, "male_hepatocytes_ageCorr_all")
save_results(female_sig, "female_hepatocytes_ageCorr_sig")
save_results(female_all, "female_hepatocytes_ageCorr_all")

# Step 4: Volcano plots
print("\n  Step 4: Creating volcano plots...")

fig, axes = plt.subplots(1, 2, figsize=(16, 7), sharey=True)
plot_volcano(male_all, "Male", axes[0])
plot_volcano(female_all, "Female", axes[1])

fig.suptitle(
    "Age-Correlated Genes in Hepatocytes\nTop 10 Up/Down Labeled per Sex",
    fontsize=16, fontweight="bold", y=1.02,
)
plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, "volcano_age_correlation.png"), dpi=300, bbox_inches="tight")
plt.savefig(os.path.join(OUTPUT_DIR, "volcano_age_correlation.pdf"), bbox_inches="tight")
plt.close()
print("  [OK] Saved: volcano_age_correlation.png/pdf")

# Step 5: Merge and classify
print("\n  Step 5: Merging male vs female results...")

merged = male_all[["gene", "correlation", "pval_adj"]].merge(
    female_all[["gene", "correlation", "pval_adj"]],
    on="gene", suffixes=("_male", "_female"),
)
merged["sig_male"] = merged["pval_adj_male"] < FDR_THRESHOLD
merged["sig_female"] = merged["pval_adj_female"] < FDR_THRESHOLD
merged["group"] = merged.apply(classify_gene_group, axis=1)

print(f"  Total genes compared: {len(merged)}")
print("\n  Gene classification:")
print(merged["group"].value_counts().to_string())

merged.to_csv(os.path.join(OUTPUT_DIR, "male_vs_female_comparison.csv"), index=False)

# Step 6: Scatter plot
print("\n  Step 6: Creating sex comparison scatter plot...")

label_groups = [
    "Male Only", "Female Only", "Both Up", "Both Down",
    "Reversal M_up_F_down", "Reversal M_down_F_up",
]

label_genes = []
for grp in label_groups:
    subset = merged[merged["group"] == grp].copy()
    if len(subset) > 0:
        subset["rank_score"] = (
            np.abs(subset["correlation_male"]) + np.abs(subset["correlation_female"])
        )
        label_genes.append(subset.nlargest(10, "rank_score"))

label_df = pd.concat(label_genes) if label_genes else pd.DataFrame()

plt.figure(figsize=(12, 12))
sns.scatterplot(
    data=merged, x="correlation_male", y="correlation_female",
    hue="group", palette=GROUP_COLORS, s=45, alpha=0.8, edgecolor=None,
    hue_order=["Not Sig", "Male Only", "Female Only",
               "Both Up", "Both Down",
               "Reversal M_up_F_down", "Reversal M_down_F_up"],
)

plt.axvline(0, linestyle="--", color="black", linewidth=1)
plt.axhline(0, linestyle="--", color="black", linewidth=1)
lims = [-0.6, 0.6]
plt.plot(lims, lims, "k:", alpha=0.5, linewidth=0.8)

for _, row in label_df.iterrows():
    dx = 0.02 if row["correlation_male"] >= 0 else -0.02
    dy = 0.02 if row["correlation_female"] >= 0 else -0.02
    plt.annotate(
        row["gene"],
        xy=(row["correlation_male"], row["correlation_female"]),
        xytext=(row["correlation_male"] + dx, row["correlation_female"] + dy),
        textcoords="data", fontsize=8, fontweight="bold",
        arrowprops=dict(arrowstyle="-", lw=0.5, color="gray"),
    )

plt.xlabel("Male Spearman Correlation (Age)", fontsize=14)
plt.ylabel("Female Spearman Correlation (Age)", fontsize=14)
plt.title(
    "Male vs Female Age Correlation in Hepatocytes\n"
    "Top 10 Genes Labeled per Group",
    fontsize=16, fontweight="bold",
)
plt.legend(title="Group", fontsize=10, title_fontsize=12,
           loc="upper left", framealpha=0.9)
sns.despine()
plt.tight_layout()

plt.savefig(os.path.join(OUTPUT_DIR, "male_vs_female_scatter.png"), dpi=300, bbox_inches="tight")
plt.savefig(os.path.join(OUTPUT_DIR, "male_vs_female_scatter.pdf"), bbox_inches="tight")
plt.savefig(os.path.join(OUTPUT_DIR, "male_vs_female_scatter.svg"), bbox_inches="tight")
plt.close()
print("  [OK] Saved: male_vs_female_scatter.png/pdf/svg")


# ==============================================================================
# PART 2: SEX COMPARISON - FILTERED ANALYSIS
# ==============================================================================

print()
print("=" * 70)
print("PART 2: Sex comparison (filtered)")
print("=" * 70)

# Step 7: Filter by correlation threshold
print(f"\n  Step 7: Filtering by |r| > {CORR_THRESHOLD}...")

male_pos = male_sig[male_sig["correlation"] > CORR_THRESHOLD].copy()
female_pos = female_sig[female_sig["correlation"] > CORR_THRESHOLD].copy()
male_neg = male_sig[male_sig["correlation"] < -CORR_THRESHOLD].copy()
female_neg = female_sig[female_sig["correlation"] < -CORR_THRESHOLD].copy()

print(f"  Male positive (r > {CORR_THRESHOLD}): {len(male_pos)}")
print(f"  Female positive (r > {CORR_THRESHOLD}): {len(female_pos)}")
print(f"  Male negative (r < -{CORR_THRESHOLD}): {len(male_neg)}")
print(f"  Female negative (r < -{CORR_THRESHOLD}): {len(female_neg)}")

male_pos.to_csv(os.path.join(COMPARISON_DIR, "male_ageCorr_positive.csv"), index=False)
female_pos.to_csv(os.path.join(COMPARISON_DIR, "female_ageCorr_positive.csv"), index=False)
male_neg.to_csv(os.path.join(COMPARISON_DIR, "male_ageCorr_negative.csv"), index=False)
female_neg.to_csv(os.path.join(COMPARISON_DIR, "female_ageCorr_negative.csv"), index=False)

male_filtered = pd.concat([male_pos, male_neg], ignore_index=True)
female_filtered = pd.concat([female_pos, female_neg], ignore_index=True)

merged_filtered = male_filtered[["gene", "correlation", "pval_adj"]].merge(
    female_filtered[["gene", "correlation", "pval_adj"]],
    on="gene", suffixes=("_male", "_female"),
)
merged_filtered["sig_male"] = merged_filtered["pval_adj_male"] < FDR_THRESHOLD
merged_filtered["sig_female"] = merged_filtered["pval_adj_female"] < FDR_THRESHOLD
merged_filtered["group"] = merged_filtered.apply(classify_gene_group, axis=1)
merged_filtered = merged_filtered[merged_filtered["group"] != "Not Sig"].copy()

print(f"\n  Filtered genes in both sexes: {len(merged_filtered)}")
print("\n  Filtered group counts:")
print(merged_filtered["group"].value_counts().to_string())

# Step 8: Scatter plots
print("\n  Step 8: Creating filtered scatter plots...")

# With group counts in legend
group_counts = merged_filtered["group"].value_counts().to_dict()
merged_filtered["group_label"] = merged_filtered["group"].apply(
    lambda g: f"{g} (n={group_counts.get(g, 0)})"
)

colors_with_counts = {
    f"{g} (n={group_counts.get(g, 0)})": GROUP_COLORS[g]
    for g in GROUP_COLORS if g in group_counts
}

labels_filtered = []
for grp in [k for k in GROUP_COLORS if k != "Not Sig"]:
    sub = merged_filtered[merged_filtered["group"] == grp].copy()
    if len(sub) > 0:
        sub["rank_score"] = np.abs(sub["correlation_male"]) + np.abs(sub["correlation_female"])
        labels_filtered.append(sub.nlargest(10, "rank_score"))

label_df_filtered = pd.concat(labels_filtered) if labels_filtered else pd.DataFrame()

plt.figure(figsize=(10, 10))
sns.scatterplot(
    data=merged_filtered, x="correlation_male", y="correlation_female",
    hue="group_label", palette=colors_with_counts, s=40, alpha=0.8,
)
plt.axvline(0, linestyle="--", color="black")
plt.axhline(0, linestyle="--", color="black")

for _, row in label_df_filtered.iterrows():
    plt.text(row["correlation_male"], row["correlation_female"],
             row["gene"], fontsize=9, weight="bold")

plt.xlabel("Male Age Correlation", fontsize=14)
plt.ylabel("Female Age Correlation", fontsize=14)
plt.title(f"Sex-Divergent Age Associations (|r| > {CORR_THRESHOLD})",
          fontsize=16, weight="bold")
plt.legend(title="Group", fontsize=10, loc="upper left")
plt.tight_layout()

plt.savefig(os.path.join(COMPARISON_DIR, "scatter_filtered_with_counts.pdf"),
            dpi=600, bbox_inches="tight")
plt.savefig(os.path.join(COMPARISON_DIR, "scatter_filtered_with_counts.png"),
            dpi=300, bbox_inches="tight")
plt.close()
print("  [OK] Saved: scatter_filtered_with_counts.pdf/png")

# Clean version (no labels)
plt.figure(figsize=(10, 10))
sns.scatterplot(
    data=merged_filtered, x="correlation_male", y="correlation_female",
    hue="group", palette={k: v for k, v in GROUP_COLORS.items() if k != "Not Sig"},
    s=45, alpha=0.8, edgecolor=None,
)
plt.axvline(0, linestyle="--", color="black")
plt.axhline(0, linestyle="--", color="black")
plt.xlabel("Male Spearman Correlation (Age)", fontsize=14)
plt.ylabel("Female Spearman Correlation (Age)", fontsize=14)
plt.title("Male vs Female Age Correlation in Hepatocytes", fontsize=16, weight="bold")
plt.legend(title="Group", fontsize=10, title_fontsize=12, loc="upper left")
sns.despine()
plt.tight_layout()
plt.savefig(os.path.join(COMPARISON_DIR, "scatter_clean.png"), dpi=300, bbox_inches="tight")
plt.savefig(os.path.join(COMPARISON_DIR, "scatter_clean.pdf"), bbox_inches="tight")
plt.close()
print("  [OK] Saved: scatter_clean.png/pdf")

# Small no-legend version
plt.figure(figsize=(5, 4))
sns.scatterplot(
    data=merged_filtered, x="correlation_male", y="correlation_female",
    hue="group", palette={k: v for k, v in GROUP_COLORS.items() if k != "Not Sig"},
    s=40, alpha=0.8, edgecolor=None, legend=False,
)
plt.axvline(0, linestyle="--", color="black", linewidth=1)
plt.axhline(0, linestyle="--", color="black", linewidth=1)
plt.xlabel("Male Spearman Correlation (Age)", fontsize=12)
plt.ylabel("Female Spearman Correlation (Age)", fontsize=12)
plt.title(f"Sex-Divergent Age Associations (|r| > {CORR_THRESHOLD})",
          fontsize=12, weight="bold")
plt.tight_layout()
plt.savefig(os.path.join(COMPARISON_DIR, "scatter_small_nolegend.png"),
            dpi=300, bbox_inches="tight")
plt.savefig(os.path.join(COMPARISON_DIR, "scatter_small_nolegend.pdf"),
            bbox_inches="tight")
plt.close()
print("  [OK] Saved: scatter_small_nolegend.png/pdf")

# Step 9: Save gene lists
print("\n  Step 9: Saving gene lists per group...")

male_only_genes = merged_filtered[merged_filtered["group"] == "Male Only"]["gene"].unique().tolist()
female_only_genes = merged_filtered[merged_filtered["group"] == "Female Only"]["gene"].unique().tolist()
both_up_genes = merged_filtered[merged_filtered["group"] == "Both Up"]["gene"].unique().tolist()
both_down_genes = merged_filtered[merged_filtered["group"] == "Both Down"]["gene"].unique().tolist()
rev_mup_fdown_genes = merged_filtered[merged_filtered["group"] == "Reversal M_up_F_down"]["gene"].unique().tolist()
rev_mdown_fup_genes = merged_filtered[merged_filtered["group"] == "Reversal M_down_F_up"]["gene"].unique().tolist()

sex_dimorphic_genes = sorted(
    male_only_genes + female_only_genes + rev_mup_fdown_genes + rev_mdown_fup_genes
)

gene_lists = {
    "Male_Only_Genes": male_only_genes,
    "Female_Only_Genes": female_only_genes,
    "Both_Up_Genes": both_up_genes,
    "Both_Down_Genes": both_down_genes,
    "Reversal_Mup_Fdown": rev_mup_fdown_genes,
    "Reversal_Mdown_Fup": rev_mdown_fup_genes,
    "Sex_Dimorphic_Genes": sex_dimorphic_genes,
}

for name, genes in gene_lists.items():
    filepath = os.path.join(COMPARISON_DIR, f"{name}.csv")
    pd.DataFrame({"gene": genes}).to_csv(filepath, index=False)
    print(f"    {name}: {len(genes)} genes")

merged_filtered.to_csv(os.path.join(COMPARISON_DIR, "merged_comparison_filtered.csv"), index=False)
merged.to_csv(os.path.join(COMPARISON_DIR, "merged_comparison_all.csv"), index=False)

# Step 10: Highlight plots for reversal patterns
print("\n  Step 10: Creating highlight plots...")

plot_highlight(
    merged_filtered, "Reversal M_up_F_down", "skyblue",
    os.path.join(COMPARISON_DIR, "highlight_reversal_Mup_Fdown.png"),
)
plot_highlight(
    merged_filtered, "Reversal M_down_F_up", "lightpink",
    os.path.join(COMPARISON_DIR, "highlight_reversal_Mdown_Fup.png"),
)


# ==============================================================================
# PART 3: PATHWAY ENRICHMENT
# ==============================================================================

print()
print("=" * 70)
print("PART 3: Pathway enrichment")
print("=" * 70)

# Step 11: Background genes
print("\n  Step 11: Preparing background genes...")

expr = adata_hep.X
gene_names = adata_hep.var_names
avg_expr = np.asarray(expr.mean(axis=0)).flatten()

expressed_genes = gene_names[avg_expr > EXPRESSION_THRESHOLD].tolist()
expressed_genes = [g.upper() for g in expressed_genes if isinstance(g, str)]

print(f"  Background genes: {len(expressed_genes)}")

# Step 12: Convert mouse to human
print("\n  Step 12: Converting mouse to human orthologs...")

mouse_gene_lists = {}
human_gene_lists = {}

gene_file_map = {
    "Male_Only": "Male_Only_Genes",
    "Female_Only": "Female_Only_Genes",
    "Both_Up": "Both_Up_Genes",
    "Both_Down": "Both_Down_Genes",
    "Reversal_Mup_Fdown": "Reversal_Mup_Fdown",
    "Reversal_Mdown_Fup": "Reversal_Mdown_Fup",
}

for group_name, file_key in gene_file_map.items():
    filepath = os.path.join(COMPARISON_DIR, f"{file_key}.csv")
    if os.path.exists(filepath):
        df_genes = pd.read_csv(filepath)
        genes = df_genes["gene"].dropna().tolist()
        genes = [g for g in genes if isinstance(g, str) and g.strip()]
        mouse_gene_lists[group_name] = genes
    else:
        mouse_gene_lists[group_name] = []
        print(f"    [SKIP] {group_name}: file not found")

for group_name, mouse_genes in mouse_gene_lists.items():
    human_genes = convert_mouse_to_human(mouse_genes)
    human_gene_lists[group_name] = human_genes
    print(f"    {group_name}: {len(mouse_genes)} mouse -> {len(human_genes)} human")

# Step 13: Run enrichment
print("\n  Step 13: Running Reactome pathway enrichment...")

all_enrichment_results = {}

for group_name, human_genes in human_gene_lists.items():
    print(f"\n    Processing {group_name}...")
    df = run_enrichr_reactome(human_genes, expressed_genes)
    if not df.empty:
        all_enrichment_results[group_name] = df
        n_sig = (df["Adjusted P-value"] < 0.05).sum()
        print(f"    Total pathways: {len(df)}, Significant: {n_sig}")
    else:
        print(f"    No enrichment results")

# Step 14: Save all pathways to Excel
print("\n  Step 14: Saving pathways to Excel...")

pathways_path = os.path.join(PATHWAY_DIR, "All_Groups_Reactome_Pathways.xlsx")

with pd.ExcelWriter(pathways_path, engine="openpyxl") as writer:
    for group_name, df in all_enrichment_results.items():
        cols = [
            "Term", "P-value", "Adjusted P-value", "Odds Ratio",
            "Combined Score", "Gene_Count", "-log10(P-value)",
            "-log10(Adj P-value)", "Overlap_Genes", "Genes",
        ]
        cols_available = [c for c in cols if c in df.columns]
        df_save = df[cols_available].sort_values("P-value")
        df_save.to_excel(writer, sheet_name=group_name[:31], index=False)

format_excel(pathways_path)
print(f"  [OK] Saved: {pathways_path}")

# Step 15: Save gene lists to Excel
print("\n  Step 15: Saving gene lists to Excel...")

genes_path = os.path.join(PATHWAY_DIR, "All_Groups_Gene_Lists.xlsx")

with pd.ExcelWriter(genes_path, engine="openpyxl") as writer:
    for group_name, mouse_genes in mouse_gene_lists.items():
        human_genes = human_gene_lists.get(group_name, [])
        max_len = max(len(mouse_genes), len(human_genes))
        df_genes = pd.DataFrame({
            "Mouse_Gene": mouse_genes + [""] * (max_len - len(mouse_genes)),
            "Human_Gene": human_genes + [""] * (max_len - len(human_genes)),
        })
        df_genes.to_excel(writer, sheet_name=group_name[:31], index=False)

format_excel(genes_path)
print(f"  [OK] Saved: {genes_path}")

# Step 16: Bubble plots
print("\n  Step 16: Creating bubble plots...")

for group_name, df in all_enrichment_results.items():
    print(f"\n    {group_name}:")
    create_bubble_plot(df, group_name, PATHWAY_DIR, top_n=20)

# Step 17: Combined comparison plot
print("\n  Step 17: Creating combined comparison plot...")

combined_list = []
for group_name, df in all_enrichment_results.items():
    df_sig = df[(df["Adjusted P-value"] < 0.05) & (df["Gene_Count"] >= 3)].copy()
    if not df_sig.empty:
        top5 = df_sig.nsmallest(5, "Adjusted P-value").copy()
        top5["Group"] = group_name
        combined_list.append(top5)

if combined_list:
    combined_df = pd.concat(combined_list, ignore_index=True)
    combined_df = combined_df.sort_values("Odds Ratio", ascending=True)
    combined_df["Term"] = pd.Categorical(
        combined_df["Term"],
        categories=combined_df["Term"].unique(),
        ordered=True,
    )

    combined_group_colors = {
        "Male_Only": "#8E44AD",
        "Female_Only": "#C0392B",
        "Both_Up": "#2ECC71",
        "Both_Down": "#16A085",
        "Reversal_Mup_Fdown": "#87CEEB",
        "Reversal_Mdown_Fup": "#FFB6C1",
    }

    n_terms = len(combined_df)
    fig_height = max(10, n_terms * 0.4)

    plt.figure(figsize=(12, fig_height))
    ax = plt.gca()

    ax.hlines(
        y=combined_df["Term"], xmin=0, xmax=combined_df["Odds Ratio"],
        color="gray", linestyle="dotted", linewidth=0.8,
    )

    scatter = sns.scatterplot(
        data=combined_df, x="Odds Ratio", y="Term",
        size="Gene_Count", hue="Group", palette=combined_group_colors,
        sizes=(80, 400), edgecolor="black", linewidth=0.6, alpha=0.9,
    )

    ax.set_xlabel("Odds Ratio", fontsize=13, fontweight="bold")
    ax.set_ylabel("")
    ax.set_title(
        "Reactome Pathways - All Sex-Dimorphic Groups\n(Top 5 per Group, Adj P < 0.05)",
        fontsize=14, fontweight="bold",
    )

    handles, labels = scatter.get_legend_handles_labels()
    ax.legend(handles, labels, bbox_to_anchor=(1.02, 1), loc="upper left",
              title="Group / Gene Count")

    plt.tight_layout()
    plt.savefig(os.path.join(PATHWAY_DIR, "Combined_All_Groups_Reactome.pdf"),
                bbox_inches="tight", dpi=300)
    plt.savefig(os.path.join(PATHWAY_DIR, "Combined_All_Groups_Reactome.png"),
                bbox_inches="tight", dpi=300)
    plt.close()
    print("  [OK] Saved: Combined_All_Groups_Reactome.pdf/png")


# ==============================================================================
# SUMMARY
# ==============================================================================

print()
print("=" * 70)
print("  COMPLETE")
print("=" * 70)
print()
print("  Part 1 - Age correlation:")
print(f"    Male tested: {len(male_all)}, Significant: {len(male_sig)}")
print(f"    Female tested: {len(female_all)}, Significant: {len(female_sig)}")
print()
print("  Part 2 - Sex comparison:")
for name, genes in gene_lists.items():
    print(f"    {name}: {len(genes)}")
print()
print("  Part 3 - Pathway enrichment:")
for group, df in all_enrichment_results.items():
    sig = (df["Adjusted P-value"] < 0.05).sum()
    print(f"    {group}: {len(df)} total, {sig} significant")
print()
print(f"  Output directories:")
print(f"    {OUTPUT_DIR}/")
print(f"    {COMPARISON_DIR}/")
print(f"    {PATHWAY_DIR}/")
print("=" * 70)
